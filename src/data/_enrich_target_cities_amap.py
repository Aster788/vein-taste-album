import json
import os
import time
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from typing import Dict, Optional, Tuple

import openpyxl

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE = os.path.join(_SCRIPT_DIR, "restaurants.xlsx")


def get_amap_web_key() -> str:
    key = os.environ.get("AMAP_WEB_KEY", "").strip()
    if not key:
        raise RuntimeError(
            "Set AMAP_WEB_KEY (Gaode Web Service key) before running this script."
        )
    return key

TARGET_CITIES = {"吉隆坡", "马六甲", "重庆", "福州", "泉州", "厦门", "广州", "青岛", "苏州", "上海"}
TARGET_COLUMNS = [
    "lng",
    "lat",
    "address",
    "cuisine",
    "price_per_person",
    "score_overall",
    "hours",
    "phone",
    "map_url",
]

QUERY_CANDIDATES = ["name_zh", "name_en", "name_local", "store_slug"]


def is_blank(v) -> bool:
    return v is None or str(v).strip() == ""


def first_non_blank(row, col_map: Dict[str, int], names) -> str:
    for name in names:
        col = col_map.get(name)
        if not col:
            continue
        value = row[col - 1]
        if not is_blank(value):
            return str(value).strip()
    return ""


def http_get_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "FoodForJoy/1.0"})
    with urllib.request.urlopen(req, timeout=25) as resp:
        return json.loads(resp.read().decode("utf-8"))


def place_text(query: str, city_zh: str, city_en: str) -> Optional[dict]:
    candidates = []
    if city_zh:
        candidates.append(city_zh)
    if city_en and city_en not in candidates:
        candidates.append(city_en)
    candidates.append("")

    api_key = get_amap_web_key()
    for city in candidates:
        params = {
            "key": api_key,
            "keywords": query,
            "extensions": "all",
            "offset": "10",
            "page": "1",
        }
        if city:
            params["city"] = city
            params["citylimit"] = "true"
        else:
            params["citylimit"] = "false"

        url = "https://restapi.amap.com/v3/place/text?" + urllib.parse.urlencode(params)
        data = http_get_json(url)
        if data.get("status") == "1" and (data.get("pois") or []):
            return (data.get("pois") or [None])[0]
    return None


def place_detail(poi_id: str) -> Optional[dict]:
    if not poi_id:
        return None
    params = {"key": get_amap_web_key(), "id": poi_id, "extensions": "all"}
    url = "https://restapi.amap.com/v3/place/detail?" + urllib.parse.urlencode(params)
    data = http_get_json(url)
    if data.get("status") != "1":
        return None
    pois = data.get("pois") or []
    return pois[0] if pois else None


def parse_location(location: str) -> Tuple[Optional[float], Optional[float]]:
    if not location or "," not in str(location):
        return None, None
    lng_raw, lat_raw = str(location).split(",", 1)
    try:
        return float(lng_raw), float(lat_raw)
    except ValueError:
        return None, None


def build_hours(detail: dict, poi: dict) -> str:
    for key in ("opentime2", "opentime"):
        v = (detail.get("biz_ext") or {}).get(key)
        if not is_blank(v):
            return str(v).strip()
    for key in ("business_hours",):
        v = detail.get(key)
        if not is_blank(v):
            return str(v).strip()
    for key in ("opentime2", "opentime"):
        v = (poi.get("biz_ext") or {}).get(key)
        if not is_blank(v):
            return str(v).strip()
    return ""


def build_cuisine(detail: dict, poi: dict) -> str:
    type_text = str(detail.get("type") or poi.get("type") or "").strip()
    if not type_text:
        return ""
    segments = [s.strip() for s in type_text.split(";") if s.strip()]
    if segments:
        return segments[-1]
    return type_text


def get_price(detail: dict, poi: dict):
    for source in (detail, poi):
        v = (source.get("biz_ext") or {}).get("cost")
        if is_blank(v):
            continue
        if isinstance(v, list):
            if not v:
                continue
            v = v[0]
        try:
            return float(v)
        except (ValueError, TypeError):
            continue
    return None


def get_score(detail: dict, poi: dict):
    for source in (detail, poi):
        v = (source.get("biz_ext") or {}).get("rating")
        if is_blank(v):
            continue
        if isinstance(v, list):
            if not v:
                continue
            v = v[0]
        try:
            return round(float(v), 1)
        except (ValueError, TypeError):
            continue
    return None


def get_phone(detail: dict, poi: dict) -> str:
    for source in (detail, poi):
        v = source.get("tel")
        if not is_blank(v):
            return str(v).strip()
    return ""


def fill_blank_cell(ws, row_idx: int, col_idx: int, value, protect_counter: Counter):
    cell = ws.cell(row_idx, col_idx)
    if not is_blank(cell.value):
        protect_counter["protected"] += 1
        return False
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return False
    cell.value = value
    return True


def main():
    wb = openpyxl.load_workbook(FILE)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    missing_required = [h for h in ["city_zh", "city_en", "name_zh", "store_slug"] + TARGET_COLUMNS if h not in col]
    if missing_required:
        raise RuntimeError(f"Missing required columns: {missing_required}")

    stats = Counter()
    protected_by_col = Counter()
    written_by_col = Counter()
    no_match_rows = []
    processed_rows = []
    row_logs = []

    for r in range(2, ws.max_row + 1):
        city_zh = str(ws.cell(r, col["city_zh"]).value or "").strip()
        if city_zh not in TARGET_CITIES:
            continue

        city_en = str(ws.cell(r, col["city_en"]).value or "").strip()
        row_values = [ws.cell(r, i + 1).value for i in range(len(headers))]
        query = first_non_blank(row_values, col, QUERY_CANDIDATES)
        if not query:
            stats["skip_no_query"] += 1
            processed_rows.append(r)
            continue

        stats["target_rows"] += 1
        processed_rows.append(r)
        poi = None
        detail = None
        try:
            poi = place_text(query, city_zh, city_en)
            if poi:
                detail = place_detail(str(poi.get("id") or "").strip()) or poi
        except Exception as e:
            no_match_rows.append({"row": r, "query": query, "city_zh": city_zh, "reason": f"request_error: {e}"})
            time.sleep(0.15)
            continue

        if not poi:
            stats["no_match"] += 1
            no_match_rows.append({"row": r, "query": query, "city_zh": city_zh, "reason": "no_poi"})
            time.sleep(0.15)
            continue

        detail = detail or poi
        lng, lat = parse_location(str(detail.get("location") or poi.get("location") or ""))
        address = str(detail.get("address") or poi.get("address") or "").strip()
        cuisine = build_cuisine(detail, poi)
        price = get_price(detail, poi)
        score = get_score(detail, poi)
        hours = build_hours(detail, poi)
        phone = get_phone(detail, poi)
        poi_id = str(detail.get("id") or poi.get("id") or "").strip()
        map_url = f"https://www.amap.com/place/{poi_id}" if poi_id else ""

        values = {
            "lng": lng,
            "lat": lat,
            "address": address,
            "cuisine": cuisine,
            "price_per_person": price,
            "score_overall": score,
            "hours": hours,
            "phone": phone,
            "map_url": map_url,
        }

        row_write = Counter()
        row_protect = Counter()
        for key in TARGET_COLUMNS:
            cell = ws.cell(r, col[key])
            if not is_blank(cell.value):
                row_protect[key] += 1
                protected_by_col[key] += 1
                continue
            v = values[key]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            cell.value = v
            written_by_col[key] += 1
            row_write[key] += 1

        if not is_blank(ws.cell(r, col["map_url"]).value) and is_blank(ws.cell(r, col["map_platform"]).value):
            ws.cell(r, col["map_platform"]).value = "amap"
            written_by_col["map_platform"] += 1

        row_logs.append(
            {
                "row": r,
                "city_zh": city_zh,
                "query": query,
                "matched_name": str(detail.get("name") or poi.get("name") or ""),
                "poi_id": poi_id,
                "written": dict(row_write),
                "protected_existing": dict(row_protect),
            }
        )
        stats["matched"] += 1
        time.sleep(0.15)

    wb.save(FILE)

    summary = {
        "file": FILE,
        "target_cities": sorted(TARGET_CITIES),
        "processed_target_rows": stats["target_rows"],
        "matched_rows": stats["matched"],
        "no_match_rows": stats["no_match"],
        "skip_no_query_rows": stats["skip_no_query"],
        "written_by_column": dict(written_by_col),
        "protected_existing_by_column": dict(protected_by_col),
        "no_match_details": no_match_rows[:200],
    }
    out_path = os.path.join(os.path.dirname(FILE), "amap_enrich_summary.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"summary_saved={out_path}")


if __name__ == "__main__":
    main()
